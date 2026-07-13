#!/usr/bin/env python3
"""
Extract tables with titles and footnotes from a PDF and save each as an
individual Excel (.xlsx) file.

Key observations for this document type:
- Tables can span multiple pages; the same title is repeated on every page.
- The header rows (column labels) also repeat on every page.
- Footnotes / abbreviations are embedded inside the table as the last rows,
  separated from the data rows by a long-underscore row ("_____...").
- A page footer ("Product Name  X von N") appears below the table's bbox
  and is ignored.

Usage:
    python3 extract_tables.py [PDF_PATH] [OUTPUT_DIR]

Defaults:
    PDF_PATH   = first *.pdf found in the current directory
    OUTPUT_DIR = extracted_tables/
"""

import sys
import re
import csv
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------

# Matches a table title line, e.g. "Table 301.1.1001.1: Summary of …"
TITLE_RE = re.compile(
    r"((?:Table|Tabelle|Liste|Figure|Abbildung)\s+[\w\.\-]+\s*[:\-]\s*.+)",
    re.IGNORECASE,
)

# Matches a separator row: first cell is all underscores (5+)
SEPARATOR_CELL_RE = re.compile(r"^_{5,}")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class TableRecord:
    title: str
    rows: List[List[str]] = field(default_factory=list)   # data rows only
    footnotes: List[str] = field(default_factory=list)    # footnote lines
    page_start: int = 0
    page_end: int = 0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def clean_cell(value) -> str:
    """Normalise a single PDF table cell to a clean string."""
    if value is None:
        return ""
    return str(value).strip()


def split_rows(raw_rows: List[List]) -> Tuple[List[List[str]], List[str]]:
    """
    Split raw table rows (as returned by pdfplumber) into:
      - data_rows   : rows before the first separator row
      - footnotes   : text lines from rows after the separator

    The separator row has its first cell filled with underscores.
    Footnote rows typically have content only in the first cell.

    Leading whitespace (indentation) in the first column is preserved for
    data rows so that hierarchical row labels keep their visual depth.
    """
    data_rows: List[List[str]] = []
    footnote_lines: List[str] = []
    past_separator = False

    for row in raw_rows:
        # Fully-stripped version of every cell – used for separator
        # detection and for footnote text (all columns).
        cleaned = [clean_cell(c) for c in row]
        first_cell = cleaned[0] if cleaned else ""

        if not past_separator and SEPARATOR_CELL_RE.match(first_cell):
            past_separator = True
            continue  # skip the separator row itself

        if past_separator:
            # Join non-empty cells in this row as one footnote line
            text = " ".join(c for c in cleaned if c)
            if text:
                footnote_lines.append(text)
        else:
            data_rows.append(cleaned)

    return data_rows, footnote_lines


def skip_header_rows(new_data_rows: List[List[str]],
                     stored_rows: List[List[str]]) -> List[List[str]]:
    """
    On continuation pages the column headers are repeated.
    Find where the new rows diverge from the stored rows and return
    only the genuinely new (non-header) rows.
    """
    skip = 0
    for i, row in enumerate(new_data_rows):
        if i < len(stored_rows) and row == stored_rows[i]:
            skip = i + 1
        else:
            break
    return new_data_rows[skip:]


def extract_title(page, table_top: float) -> Optional[str]:
    """Return the last 'Table …:' title line found above the table bbox."""
    try:
        above = page.crop((0, 0, page.width, table_top))
        text = above.extract_text() or ""
    except Exception:
        return None
    matches = TITLE_RE.findall(text)
    return matches[-1].strip() if matches else None


# ---------------------------------------------------------------------------
# Indentation detection from cell x-positions
# ---------------------------------------------------------------------------

def _compute_indent_unit(page, tobj) -> float:
    """
    Scan every first-column cell in *tobj* and collect the x-offset of the
    first word from the cell's left edge.  The smallest positive offset
    found is returned as the width of one indent level (PDF points).
    Falls back to 10 pt when no indented rows are present.
    """
    positive: List[float] = []
    for row_obj in tobj.rows:
        cells = getattr(row_obj, "cells", None) or []
        if not cells or cells[0] is None:
            continue
        bbox = cells[0]
        try:
            words = page.within_bbox(bbox, relative=False).extract_words(
                x_tolerance=3, y_tolerance=3
            )
        except Exception:
            continue
        if not words:
            continue
        offset = float(words[0]["x0"]) - float(bbox[0])
        if offset > 2.0:          # ignore floating-point noise near zero
            positive.append(offset)
    return min(positive) if positive else 10.0


def _get_indent_str(page, cell_bbox, indent_unit: float, tolerance: float) -> str:
    """
    Return a string of leading spaces for the row label in *cell_bbox*.

    Indentation is inferred from how far the first word sits from the
    cell's left edge, expressed as a multiple of *indent_unit*.  Only
    "clean" multiples (within *tolerance* of an exact multiple) are
    accepted; off-grid offsets – typical of centred column-header text –
    are silently treated as zero indentation.
    """
    if cell_bbox is None:
        return ""
    try:
        words = page.within_bbox(cell_bbox, relative=False).extract_words(
            x_tolerance=3, y_tolerance=3
        )
    except Exception:
        return ""
    if not words:
        return ""
    offset = max(0.0, float(words[0]["x0"]) - float(cell_bbox[0]))
    if offset < 2.0:              # essentially zero
        return ""
    level = round(offset / indent_unit)
    if level == 0:
        return ""
    # Reject if the offset does not sit close to a clean multiple of
    # indent_unit – this filters out centred headers whose x-position
    # accidentally resembles an indent level.
    if abs(offset - level * indent_unit) > tolerance:
        return ""
    return "  " * level            # 2 spaces per indent level


def extract_rows_with_indent(
    tobj,
    page,
) -> Tuple[List[List[str]], List[str]]:
    """
    Extract table text rows and footnotes from *tobj*, attaching
    first-column indentation measured from word x-positions in the cell
    bounding boxes.

    This replaces the plain ``split_rows(tobj.extract())`` call so that
    hierarchical row labels keep their visual indent depth.
    """
    raw_rows = tobj.extract()
    if not raw_rows:
        return [], []

    # Per-table indent calibration
    indent_unit = _compute_indent_unit(page, tobj)
    tolerance   = 0.25 * indent_unit   # tight – rejects off-grid header text

    # First-column cell bboxes, parallel to raw_rows / tobj.rows
    first_col_bboxes: List[Optional[Tuple]] = []
    for row_obj in tobj.rows:
        cells = getattr(row_obj, "cells", None) or []
        first_col_bboxes.append(cells[0] if cells else None)

    data_rows:      List[List[str]] = []
    footnote_lines: List[str]       = []
    past_separator = False

    for idx, raw_row in enumerate(raw_rows):
        cleaned    = [clean_cell(c) for c in raw_row]
        first_cell = cleaned[0] if cleaned else ""

        if not past_separator and SEPARATOR_CELL_RE.match(first_cell):
            past_separator = True
            continue

        if past_separator:
            text = " ".join(c for c in cleaned if c)
            if text:
                footnote_lines.append(text)
        else:
            bbox   = first_col_bboxes[idx] if idx < len(first_col_bboxes) else None
            indent = _get_indent_str(page, bbox, indent_unit, tolerance)
            data_rows.append([indent + cleaned[0]] + cleaned[1:])

    return data_rows, footnote_lines


def sanitize_filename(title: str, max_len: int = 160) -> str:
    """Turn a table title into a safe filesystem name."""
    name = re.sub(r'[<>:"/\\|?*\r\n\t]', "_", title)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:max_len]


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(record: TableRecord, path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table"

    # Styles
    title_font  = Font(bold=True, size=12)
    hdr_font    = Font(bold=True, size=10)
    hdr_fill    = PatternFill("solid", fgColor="D9E1F2")
    data_font   = Font(size=10)
    fn_label    = Font(bold=True, size=9)
    fn_font     = Font(italic=True, size=9, color="595959")
    thin        = Side(style="thin")
    bdr         = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top    = Alignment(wrap_text=True, vertical="top")

    num_cols = max((len(r) for r in record.rows), default=1)
    span     = max(num_cols, 2)

    row = 1

    # ---- Title ----
    if record.title:
        cell = ws.cell(row=row, column=1, value=record.title)
        cell.font = title_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=span)
        ws.row_dimensions[row].height = 32
        row += 2  # title + blank line

    # ---- Table data ----
    for r_idx, data_row in enumerate(record.rows):
        for c_idx in range(num_cols):
            val  = data_row[c_idx] if c_idx < len(data_row) else ""
            cell = ws.cell(row=row + r_idx, column=c_idx + 1, value=val)
            cell.border    = bdr
            cell.alignment = wrap_top
            cell.font      = hdr_font if r_idx == 0 else data_font
            if r_idx == 0:
                cell.fill = hdr_fill

    row += len(record.rows)

    # ---- Footnotes ----
    if record.footnotes:
        row += 1  # blank row between data and footnote area

        # Separator row – underscores spanning the full table width.
        # Written to the output so that downstream readers can reliably
        # detect the boundary between data rows and footnotes.
        sep_cell = ws.cell(row=row, column=1, value="_" * 40)
        sep_cell.font      = Font(size=8, color="888888")
        sep_cell.alignment = Alignment(horizontal="left")
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=span)
        row += 1

        # "Notes / Abbreviations:" label – merged across full table width
        lbl = ws.cell(row=row, column=1, value="Notes / Abbreviations:")
        lbl.font      = fn_label
        lbl.alignment = Alignment(wrap_text=True)
        if span > 1:
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=span)
        row += 1

        # Each footnote line – merged across the full table width
        for line in record.footnotes:
            cell = ws.cell(row=row, column=1, value=line)
            cell.font      = fn_font
            cell.alignment = Alignment(wrap_text=True)
            if span > 1:
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=span)
            row += 1

    # ---- Column widths ----
    for col_idx in range(1, num_cols + 1):
        lengths = [
            len(r[col_idx - 1])
            for r in record.rows
            if col_idx - 1 < len(r) and r[col_idx - 1]
        ]
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            min(max(lengths, default=10) + 2, 55)
        )

    wb.save(path)


# ---------------------------------------------------------------------------
# Main extraction logic
# ---------------------------------------------------------------------------

def extract_tables(pdf_path: Path, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    records: List[TableRecord] = []
    current: Optional[TableRecord] = None

    print(f"Opening: {pdf_path}")

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        print(f"Pages: {total}")

        for page_num, page in enumerate(pdf.pages, start=1):
            if page_num % 200 == 0:
                print(f"  ... page {page_num}/{total}")

            table_objects = page.find_tables()
            if not table_objects:
                continue

            for tobj in table_objects:
                data_rows, footnotes = extract_rows_with_indent(tobj, page)
                if not data_rows:
                    continue

                title = extract_title(page, tobj.bbox[1])

                # ---- continuation or new table? ----
                same_title = (
                    title is not None
                    and current is not None
                    and current.title == title
                )
                same_columns = (
                    title is None
                    and current is not None
                    and data_rows
                    and current.rows
                    and data_rows[0] == current.rows[0]
                )
                is_continuation = same_title or same_columns

                if is_continuation:
                    new_rows = skip_header_rows(data_rows, current.rows)
                    current.rows.extend(new_rows)
                    current.page_end = page_num
                    # Overwrite footnotes: last page has the most complete text
                    if footnotes:
                        current.footnotes = footnotes
                else:
                    if current is not None:
                        records.append(current)

                    current = TableRecord(
                        title=title or f"Table_page{page_num}",
                        rows=data_rows,
                        footnotes=footnotes,
                        page_start=page_num,
                        page_end=page_num,
                    )

        if current is not None:
            records.append(current)

    print(f"\nFound {len(records)} unique tables.")

    # ---- Save ----
    for idx, rec in enumerate(records, start=1):
        filename  = f"{idx:04d}_{sanitize_filename(rec.title)}.xlsx"
        out_path  = output_dir / filename
        fn_info   = f", {len(rec.footnotes)} footnote line(s)" if rec.footnotes else ""
        try:
            write_excel(rec, out_path)
            print(
                f"  [{idx:4d}/{len(records)}] "
                f"p{rec.page_start}–{rec.page_end}  "
                f"{len(rec.rows)} rows{fn_info}  →  {filename}"
            )
        except Exception as exc:
            print(f"  ERROR [{idx}] {filename}: {exc}")

    # ---- Page index (enables PDF vs RTF comparison in rtf_tables_app.R) ----
    index_path = output_dir / "page_index.csv"
    try:
        with open(index_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["nr", "page_start", "page_end", "filename"])
            for idx, rec in enumerate(records, start=1):
                filename = f"{idx:04d}_{sanitize_filename(rec.title)}.xlsx"
                writer.writerow([idx, rec.page_start, rec.page_end, filename])
        print(f"Page index:  {index_path.resolve()}")
    except Exception as exc:
        print(f"  Warning: could not write page index: {exc}")

    print(f"\nDone. Output: {output_dir.resolve()}/")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) >= 2:
        pdf_path = Path(sys.argv[1])
    else:
        pdfs = list(Path(".").glob("*.pdf"))
        if not pdfs:
            sys.exit("No PDF file found. Pass the path as the first argument.")
        pdf_path = pdfs[0]

    if not pdf_path.exists():
        sys.exit(f"File not found: {pdf_path}")

    output_dir = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path("extracted_tables")
    extract_tables(pdf_path, output_dir)


if __name__ == "__main__":
    main()
